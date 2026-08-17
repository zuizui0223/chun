#!/usr/bin/env python3
"""Audit Fan2026 Supplementary Table S2 anthocyanin matrix.

Crosswalk Table S2 quantitative anthocyanin rows to Table S1 petal-colour
records from the same workbook. Exact/source matches are admitted first;
species consensus is a conservative fallback. Unresolved rows receive ranked
same-workbook spelling candidates for audit, but fuzzy matches are never
silently admitted as biological identities.
"""
from __future__ import annotations
import argparse,csv,json,math,re
from collections import defaultdict
from difflib import SequenceMatcher
from itertools import combinations
from pathlib import Path
from statistics import mean,median,stdev
from openpyxl import load_workbook

def fnum(x):
    try:return float(x or 0)
    except Exception:return 0.0

def clean(x):return re.sub(r"\s+"," ",str(x or "").replace("\xa0"," ").strip())

def taxon_unit(s):
    s=clean(s);s=re.sub(r"[_-]\d+$","",s)
    if s.startswith("C. "):s="Camellia "+s[3:]
    elif s.startswith("C."):s="Camellia "+s[2:].lstrip()
    return s

def species_base(s):
    s=taxon_unit(s);s=re.sub(r"'[^']+'$","",s).strip()
    s=re.split(r"\s+(?:var\.?|subsp\.?|ssp\.?|f\.?|forma|from\.)\s+",s,maxsplit=1,flags=re.I)[0]
    p=s.split();return " ".join(p[:2]) if len(p)>=2 and p[0]=="Camellia" else ""

def state(x):
    x=clean(x).lower()
    if x=="white":return "W"
    if x=="yellow":return "Y"
    if x in {"red","purple red","pink","plink"}:return "A"
    return ""

def bray_comp(a,b):
    sa=sum(a);sb=sum(b)
    if sa<=0 or sb<=0:return None
    return 0.5*sum(abs(x/sa-y/sb) for x,y in zip(a,b))

def q(xs,p):
    ys=sorted(xs);n=len(ys)
    if not n:return None
    h=(n-1)*p;i=math.floor(h);j=math.ceil(h)
    return ys[i] if i==j else ys[i]*(j-h)+ys[j]*(h-i)

def state_stats(rows,st):
    rr=[r for r in rows if r['colour_state']==st]
    if not rr:return {f'{st}_n':0}
    vals=[r['total_reported_anthocyanin'] for r in rr];rich=[r['detected_anthocyanin_richness'] for r in rr]
    return {f'{st}_n':len(rr),f'{st}_anthocyanin_min':min(vals),f'{st}_anthocyanin_median':median(vals),f'{st}_anthocyanin_max':max(vals),f'{st}_anthocyanin_max_to_min_ratio':max(vals)/min(vals) if min(vals)>0 else None,f'{st}_richness_median':median(rich)}

def similarity(a,b):return SequenceMatcher(None,a.lower(),b.lower()).ratio()

def main():
    ap=argparse.ArgumentParser();ap.add_argument('--workbook',type=Path,required=True);ap.add_argument('--output-dir',type=Path,required=True);ap.add_argument('--article-expected-red-like-species',type=int,default=29);a=ap.parse_args()
    wb=load_workbook(a.workbook,read_only=True,data_only=True)
    s1=wb['Table.S1'];s1rows=list(s1.iter_rows(values_only=True));hi=None;header=None
    for i,r in enumerate(s1rows):
        vals=[clean(x) for x in r]
        if {'Species','Petal color'}.issubset(set(vals)):hi=i;header=vals;break
    if hi is None:raise SystemExit('Table S1 petal-colour header not found')
    idx={h:j for j,h in enumerate(header) if h}
    exact=defaultdict(lambda:{'states':set(),'colours':set(),'sections':set()});by_species=defaultdict(lambda:{'states':set(),'colours':set(),'sections':set()})
    for r in s1rows[hi+1:]:
        vals=[clean(x) for x in r]
        if not any(vals):continue
        if 'Genus' in idx and vals[idx['Genus']]!='Camellia':continue
        u=taxon_unit(vals[idx['Species']]);sp=species_base(u);st=state(vals[idx['Petal color']])
        if not u or not st:continue
        sec=vals[idx['Section']] if 'Section' in idx else ''
        for d in [exact[u],by_species[sp]]:
            d['states'].add(st);d['colours'].add(vals[idx['Petal color']].lower())
            if sec:d['sections'].add(sec)
    exact_keys=sorted(exact)

    ws=wb['Table.S2'];headers=[clean(x) for x in next(ws.iter_rows(min_row=2,max_row=2,values_only=True))];pigment_headers=headers[1:24]
    rows=[];matrix=[]
    for r in ws.iter_rows(min_row=3,values_only=True):
        name=clean(r[0])
        if not name or name.lower().startswith('note'):break
        vals=[fnum(x) for x in r[1:24]];matrix.append(vals)
        total=sum(vals);rich=sum(x>0 for x in vals);cy=sum(vals[i] for i,h in enumerate(pigment_headers) if h.startswith('Cy'));dp=sum(vals[i] for i,h in enumerate(pigment_headers) if h.startswith('Dp'));frac=cy/total if total>0 else None;p=[x/total for x in vals if x>0] if total>0 else [];sh=-sum(x*math.log(x) for x in p) if p else None
        u=taxon_unit(name);sp=species_base(u);match='unresolved';d=None
        if u in exact and len(exact[u]['states'])==1:d=exact[u];match='exact_taxon'
        elif sp in by_species and len(by_species[sp]['states'])==1:d=by_species[sp];match='species_consensus'
        st=next(iter(d['states'])) if d else ''
        cand='';cand_score='';cand_state='';cand_colour=''
        if not d:
            ranked=sorted(((similarity(u,k),k) for k in exact_keys),reverse=True)[:3]
            cand=' | '.join(k for _,k in ranked);cand_score=' | '.join(f'{s:.4f}' for s,_ in ranked)
            if ranked:
                bd=exact[ranked[0][1]];cand_state=';'.join(sorted(bd['states']));cand_colour=';'.join(sorted(bd['colours']))
        rows.append({'species_label':name,'normalized_taxon':u,'species_base':sp,'colour_state':st,'source_petall_colour':';'.join(sorted(d['colours'])) if d else '','source_section':';'.join(sorted(d['sections'])) if d else '','colour_crosswalk_status':match,'closest_s1_taxa':cand,'closest_s1_similarity':cand_score,'closest_s1_top_state':cand_state,'closest_s1_top_colour':cand_colour,'total_reported_anthocyanin':total,'detected_anthocyanin_richness':rich,'cyanidin_total':cy,'delphinidin_total':dp,'cyanidin_fraction':frac,'composition_shannon':sh,'claim_ceiling':'fuzzy candidates are audit hints only; biological state is admitted only from exact/same-species consensus'})
    if not rows:raise SystemExit('no Table S2 species rows found')
    dists=[bray_comp(matrix[i],matrix[j]) for i,j in combinations(range(len(matrix)),2)];dists=[x for x in dists if x is not None];totals=[r['total_reported_anthocyanin'] for r in rows];rich=[r['detected_anthocyanin_richness'] for r in rows];cf=[r['cyanidin_fraction'] for r in rows if r['cyanidin_fraction'] is not None]
    byname={r['species_label']:k for k,r in enumerate(rows)};jp=next((k for n,k in byname.items() if n=='C. japonica'),None);ru=next((k for n,k in byname.items() if 'rusticana' in n),None);pair={}
    if jp is not None and ru is not None:pair={'C_japonica_vs_rusticana_composition_bray':bray_comp(matrix[jp],matrix[ru]),'rusticana_to_japonica_total_ratio':totals[ru]/totals[jp] if totals[jp]>0 else None,'C_japonica_richness':rich[jp],'C_rusticana_richness':rich[ru]}
    cc={st:sum(r['colour_state']==st for r in rows) for st in ['A','W','Y']};unresolved=[r for r in rows if not r['colour_state']]
    summary={'source':'Fan2026_PMC12946509_TableS1_TableS2','article_expected_red_like_species':a.article_expected_red_like_species,'supplement_species_label_rows':len(rows),'colour_crosswalk_A':cc['A'],'colour_crosswalk_W':cc['W'],'colour_crosswalk_Y':cc['Y'],'colour_crosswalk_unresolved':len(unresolved),'article_red_like_count_matches_crosswalk_A':cc['A']==a.article_expected_red_like_species,'unresolved_top_candidates':[{'s2':r['species_label'],'s1':r['closest_s1_taxa'].split(' | ')[0],'score':r['closest_s1_similarity'].split(' | ')[0],'candidate_state':r['closest_s1_top_state'],'candidate_colour':r['closest_s1_top_colour']} for r in unresolved],'article_stated_richness_range':'2-15','supplement_row_richness_min':min(rich),'supplement_row_richness_median':median(rich),'supplement_row_richness_max':max(rich),'rows_below_article_stated_min_richness':sum(x<2 for x in rich),'total_anthocyanin_min':min(totals),'total_anthocyanin_median':median(totals),'total_anthocyanin_max':max(totals),'total_max_to_min_ratio':max(totals)/min(totals) if min(totals)>0 else None,'total_cv':stdev(totals)/mean(totals),'pairwise_normalized_composition_bray_q25':q(dists,.25),'pairwise_normalized_composition_bray_median':median(dists),'pairwise_normalized_composition_bray_q75':q(dists,.75),'pairwise_normalized_composition_bray_max':max(dists),'rows_cyanidin_fraction_gt_0_9':sum(x>.9 for x in cf),'rows_with_delphinidin':sum(r['delphinidin_total']>0 for r in rows),**state_stats(rows,'A'),**state_stats(rows,'W'),**state_stats(rows,'Y'),**pair,'decision':'exact/species-consensus crosswalk remains the admission gate; closest Table S1 candidates expose likely spelling mismatches for manual/source validation','claim_ceiling':'within-state quantitative heterogeneity; fuzzy taxon candidates are not admitted identities'}
    a.output_dir.mkdir(parents=True,exist_ok=True)
    with (a.output_dir/'species_row_summary.csv').open('w',newline='',encoding='utf-8') as f:w=csv.DictWriter(f,fieldnames=list(rows[0]));w.writeheader();w.writerows(rows)
    (a.output_dir/'summary.json').write_text(json.dumps(summary,indent=2)+'\n');print(json.dumps(summary,indent=2))
if __name__=='__main__':main()
