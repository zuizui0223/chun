#!/usr/bin/env python3
"""Inventory sheet names and small previews from supplementary tabular files."""
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import load_workbook


def cell_text(v):
    if v is None:
        return ""
    s = str(v).replace("\n", " ").replace("\r", " ").strip()
    return s[:160]


def inspect_xlsx(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        preview = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
            preview.append(" | ".join(cell_text(x) for x in row[: min(ws.max_column, 12)]))
        out.append({
            "file": str(path),
            "sheet": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "preview": " || ".join(preview),
        })
    return out


def inspect_delimited(path: Path):
    delim = "\t" if path.suffix.lower() == ".tsv" else ","
    rows = []
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
        reader = csv.reader(fh, delimiter=delim)
        for i, row in enumerate(reader):
            rows.append(" | ".join(cell_text(x) for x in row[:12]))
            if i >= 7:
                break
    return [{"file": str(path), "sheet": "", "max_row": "", "max_column": "", "preview": " || ".join(rows)}]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", type=Path, required=True)
    ap.add_argument("--output", type=Path, required=True)
    args = ap.parse_args()
    records = []
    failures = []
    for path in sorted(p for p in args.root.rglob("*") if p.is_file()):
        try:
            if path.suffix.lower() == ".xlsx":
                records.extend(inspect_xlsx(path))
            elif path.suffix.lower() in {".csv", ".tsv"}:
                records.extend(inspect_delimited(path))
        except Exception as exc:
            failures.append({"file": str(path), "error": f"{type(exc).__name__}: {exc}"})
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", newline="", encoding="utf-8") as fh:
        fields = ["file", "sheet", "max_row", "max_column", "preview"]
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        w.writerows(records)
    (args.output.parent / "inspection_failures.json").write_text(json.dumps(failures, indent=2) + "\n")
    print(json.dumps({"tables_or_sheets": len(records), "failures": len(failures)}, indent=2))


if __name__ == "__main__":
    main()
