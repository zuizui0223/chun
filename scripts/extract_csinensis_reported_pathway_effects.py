#!/usr/bin/env python3
"""Extract stage-matched white-vs-pink pathway effects from Zhou et al. 2020.

The supplements contain an important provenance inconsistency: Table S6 sample
headers are systematically reversed relative to Table S8, Table S9, the paper
text and the public sample mapping (ZJW=white; BTP=pink). We do not silently
correct this. The script first demonstrates the swap on overlapping gene-stage
FPKM triplets, then admits S6 only after that gate passes.

All outputs belong to one biological independence cluster (CSIN_WHITE_PINK).
Genes and stages are repeated measurements, not independent studies.
"""
from __future__ import annotations

import argparse,csv,json,math,re
from collections import defaultdict
from pathlib import Path
from statistics import mean,median,stdev
from openpyxl import load_workbook

STAGES=range(1,6)
PATTERNS=[
 ('DFR',re.compile(r'\bDFR\b|DIHYDROFLAVONOL[- ]?4[- ]?REDUCTASE',re.I)),
 ('ANS',re.compile(r'\bANS\b|ANTHOCYANIDIN SYNTHASE',re.I)),
 ('UFGT',re.compile(r'\bUFGT\b|UDP.*GLUCOSYLTRANSFERASE',re.I)),
 ('FLS',re.compile(r'\bFLS\d*\b|FLAVONOL SYNTHASE',re.I)),
 ('ANR',re.compile(r'\bANR\b|ANTHOCYANIDIN REDUCTASE',re.I)),
 ('LAR',re.compile(r'\bF?LAR\b|LEUCOANTHOCYANIDIN REDUCT',re.I)),
 ('CHS',re.compile(r'\bCHS\d*\b|CHALCONE SYNTHASE',re.I)),
 ('F3H_5PRIME',re.compile(r"F3['’]5['’]H",re.I)),
 ('F3H_PRIME',re.compile(r"F3['’]H",re.I)),
 ('F3H',re.compile(r'\bF3H\b|FLAVANONE 3[- ]HYDROXYLASE',re.I)),
]

def gclass(x):
    s=str(x or '').strip()
    if not s or s=='--':return None
    for name,p in PATTERNS:
        if p.search(s):return name
    return None

def fnum(x):
    try:
        v=float(x)
        return v if math.isfinite(v) and v>=0 else None
    except Exception:return None

def close3(a,b,tol=1e-7):
    return len(a)==len(b)==3 and all(abs(x-y)<=tol*max(1.0,abs(x),abs(y)) for x,y in zip(a,b))

def log_effect(white,pink):
    lw=[math.log2(x+1) for x in white];lp=[math.log2(x+1) for x in pink]
    return mean(lp)-mean(lw),lw,lp

def hedges_g(lp,lw):
    if len(lp)<2 or len(lw)<2:return None
    sp2=((len(lp)-1)*stdev(lp)**2+(len(lw)-1)*stdev(lw)**2)/(len(lp)+len(lw)-2)
    if sp2<=0:return None
    d=(mean(lp)-mean(lw))/math.sqrt(sp2)
    j=1-3/(4*(len(lp)+len(lw))-9)
    return j*d

def module(gc):
    if gc in {'DFR','ANS','UFGT'}:return 'anthocyanin_downstream'
    if gc=='FLS':return 'flavonol_branch'
    if gc in {'ANR','LAR'}:return 'proanthocyanidin_branch'
    if gc in {'CHS','F3H','F3H_PRIME','F3H_5PRIME'}:return 'shared_or_upstream_flavonoid'
    return None

def effect_row(gid,gc,stage,source,white,pink,reported_log2fc=None,fdr=None):
    de,lw,lp=log_effect(white,pink)
    return {'independence_cluster':'CSIN_WHITE_PINK','gene_id':gid,'gene_class':gc,'stage':stage,'source_table':source,
      'n_white':len(white),'n_pink':len(pink),'mean_white_fpkm':mean(white),'mean_pink_fpkm':mean(pink),
      'pink_minus_white_mean_log2fpkm':de,'hedges_g_log2fpkm':hedges_g(lp,lw),
      'author_reported_log2fc':reported_log2fc if reported_log2fc is not None else '',
      'author_reported_fdr':fdr if fdr is not None else '',
      'claim_ceiling':'within-study processed-expression effect; genes/stages are repeated measures, not independent studies'}

def load_s8(path):
    wb=load_workbook(path,read_only=True,data_only=True);ann={};profile={};rows=[]
    for stage in STAGES:
        ws=next(x for x in wb.worksheets if f'stage{stage}' in x.title.lower())
        for r in ws.iter_rows(values_only=True):
            gid=str(r[0] or '').strip()
            if not gid or gid.startswith('#'):continue
            gc=gclass(r[1] if len(r)>1 else None)
            if not gc:continue
            ann[gid]=gc
            if len(r)<14:continue
            white=[fnum(r[i]) for i in (2,4,6)];pink=[fnum(r[i]) for i in (8,10,12)]
            if any(x is None for x in white+pink):continue
            profile[(gid,stage)]=(white,pink)
            fdr=fnum(r[14]) if len(r)>14 else None
            lfc=float(r[15]) if len(r)>15 and r[15] is not None else None
            rows.append(effect_row(gid,gc,stage,'S8_stage_selected_DEG',white,pink,lfc,fdr))
    return ann,profile,rows

def header(ws,pred,maxr=8):
    for i,r in enumerate(ws.iter_rows(min_row=1,max_row=maxr,values_only=True),1):
        v=[str(x or '').strip() for x in r]
        if pred(v):return i,v
    return None,None

def read_s6_raw(path):
    wb=load_workbook(path,read_only=True,data_only=True);ws=wb.active
    hr,h=header(ws,lambda v:'#ID' in v and any(x.endswith('_FPKM') for x in v));idx={x:i for i,x in enumerate(h)}
    data={};ann={}
    for r in ws.iter_rows(min_row=hr+1,values_only=True):
        gid=str(r[idx['#ID']] or '').strip();gc=gclass(r[idx['gene_name']] if gid else None)
        if not gid or not gc:continue
        ann[gid]=gc
        for st in STAGES:
            lab_b=[fnum(r[idx[f'BTP{st}{q}_FPKM']]) for q in (1,2,3)]
            lab_z=[fnum(r[idx[f'ZJW{st}{q}_FPKM']]) for q in (1,2,3)]
            if not any(x is None for x in lab_b+lab_z):data[(gid,st)]=(lab_z,lab_b) # labelled ZJW, labelled BTP
    return ann,data

def audit_s6_labels(s6,s8):
    direct=swapped=overlap=0;examples=[]
    for key,(s6_z,s6_b) in s6.items():
        if key not in s8:continue
        overlap+=1;w,p=s8[key]
        d=close3(s6_z,w) and close3(s6_b,p)
        s=close3(s6_b,w) and close3(s6_z,p)
        direct+=int(d);swapped+=int(s)
        examples.append({'gene_id':key[0],'stage':key[1],'direct_label_match':d,'swapped_label_match':s})
    if overlap<3 or swapped<=direct or swapped<3:
        raise SystemExit(f'S6 label provenance gate unresolved overlap={overlap} direct={direct} swapped={swapped}')
    summary={'table':'S6','overlap_gene_stage_tests':overlap,'direct_matches':direct,'swapped_matches':swapped,
      'decision':'systematic BTP/ZJW header swap verified against S8; interpret S6 labelled BTP as biological ZJW/white and labelled ZJW as biological BTP/pink',
      'claim_ceiling':'supplement metadata correction inside Zhou2020 only'}
    return summary,examples

def s6_effects(s6,ann,swap=True):
    out=[]
    for (gid,st),(lab_z,lab_b) in sorted(s6.items(),key=lambda x:(x[0][1],x[0][0])):
        white,pink=(lab_b,lab_z) if swap else (lab_z,lab_b)
        out.append(effect_row(gid,ann[gid],st,'S6_hub_full_profile_label_swap_corrected' if swap else 'S6_hub_full_profile',white,pink))
    return out

def load_s9(path,ann,s8):
    wb=load_workbook(path,read_only=True,data_only=True);ws=wb.active
    hr,h=header(ws,lambda v:'modColor' in v and 'ZJW11' in v and 'BTP11' in v);idx={x:i for i,x in enumerate(h)}
    out=[];direct=swapped=overlap=0
    for r in ws.iter_rows(min_row=hr+1,values_only=True):
        gid=str(r[0] or '').strip()
        if gid not in ann:continue
        for st in STAGES:
            keys_z=[f'ZJW{st}{q}' for q in (1,2,3)];keys_b=[f'BTP{st}{q}' for q in (1,2,3)]
            if not all(k in idx for k in keys_z+keys_b):continue
            white=[fnum(r[idx[k]]) for k in keys_z];pink=[fnum(r[idx[k]]) for k in keys_b]
            if any(x is None for x in white+pink):continue
            if (gid,st) in s8:
                overlap+=1;w,p=s8[(gid,st)];direct+=int(close3(white,w) and close3(pink,p));swapped+=int(close3(pink,w) and close3(white,p))
            out.append(effect_row(gid,ann[gid],st,'S9_module_full_profile',white,pink))
    audit={'table':'S9','overlap_gene_stage_tests':overlap,'direct_matches':direct,'swapped_matches':swapped,
      'decision':'labels retained as printed; overlap with S8 supports direct ZJW=white, BTP=pink mapping',
      'claim_ceiling':'supplement metadata check inside Zhou2020 only'}
    return out,audit

def mod_summary(rows,source_label):
    by=defaultdict(list);out=[]
    for r in rows:
        m=module(r['gene_class'])
        if m:by[(r['stage'],m)].append(r)
    for (st,m),rs in sorted(by.items()):
        vals=[float(r['pink_minus_white_mean_log2fpkm']) for r in rs]
        out.append({'independence_cluster':'CSIN_WHITE_PINK','source_set':source_label,'stage':st,'module':m,'n_genes':len(vals),
          'mean_gene_log2_effect':mean(vals),'median_gene_log2_effect':median(vals),'n_positive_genes':sum(v>0 for v in vals),'n_negative_genes':sum(v<0 for v in vals),
          'gene_ids':';'.join(r['gene_id'] for r in rs),'claim_ceiling':'within-study selected gene/module summary; not a natural pathway-frequency estimate'})
    return out

def contrasts(ms):
    look={(r['source_set'],r['stage'],r['module']):r for r in ms};out=[]
    for src in sorted({r['source_set'] for r in ms}):
      for st in STAGES:
        a=look.get((src,st,'anthocyanin_downstream'));f=look.get((src,st,'flavonol_branch'));p=look.get((src,st,'proanthocyanidin_branch'))
        for name,x in [('anthocyanin_minus_flavonol',f),('anthocyanin_minus_proanthocyanidin',p)]:
          if a and x:out.append({'independence_cluster':'CSIN_WHITE_PINK','source_set':src,'stage':st,'contrast':name,'difference_of_module_log2_effects':float(a['mean_gene_log2_effect'])-float(x['mean_gene_log2_effect']),
            'claim_ceiling':'descriptive relative allocation contrast; PA kept distinct from FLS by design'})
    return out

def write(path,rows):
    path.parent.mkdir(parents=True,exist_ok=True);fields=list(rows[0]) if rows else ['empty']
    with path.open('w',newline='',encoding='utf-8') as f:
        w=csv.DictWriter(f,fieldnames=fields);w.writeheader();w.writerows(rows)

def main():
    ap=argparse.ArgumentParser();ap.add_argument('--s6',type=Path,required=True);ap.add_argument('--s8',type=Path,required=True);ap.add_argument('--s9',type=Path,required=True);ap.add_argument('--out-dir',type=Path,required=True);a=ap.parse_args()
    ann8,p8,r8=load_s8(a.s8);ann6,s6=read_s6_raw(a.s6);ann={**ann8,**ann6}
    s6audit,examples=audit_s6_labels(s6,p8);r6=s6_effects(s6,ann6,True);r9,s9audit=load_s9(a.s9,ann,p8)
    m6=mod_summary(r6,'S6_hub_full_profile_label_swap_corrected');m8=mod_summary(r8,'S8_stage_selected_DEG');m9=mod_summary(r9,'S9_module_full_profile')
    cs=contrasts(m6+m8+m9)
    write(a.out_dir/'s6_hub_gene_stage_effects.csv',r6);write(a.out_dir/'s8_selected_deg_effects.csv',r8);write(a.out_dir/'s9_module_gene_stage_effects.csv',r9)
    write(a.out_dir/'module_stage_effects.csv',m6+m8+m9);write(a.out_dir/'module_contrasts.csv',cs);write(a.out_dir/'s6_label_swap_examples.csv',examples)
    (a.out_dir/'provenance_label_audit.json').write_text(json.dumps({'S6':s6audit,'S9':s9audit},indent=2)+'\n')
    dfr6=[r for r in r6 if r['gene_class']=='DFR'];fls6=[r for r in r6 if r['gene_class']=='FLS'];lar6=[r for r in r6 if r['gene_class']=='LAR']
    c6=[r for r in cs if r['source_set']=='S6_hub_full_profile_label_swap_corrected']
    summary={'independence_cluster':'CSIN_WHITE_PINK','s6_label_gate':s6audit,'s9_label_gate':s9audit,
      'S6_DFR_positive_gene_stage':f"{sum(float(r['pink_minus_white_mean_log2fpkm'])>0 for r in dfr6)}/{len(dfr6)}",
      'S6_FLS_positive_gene_stage':f"{sum(float(r['pink_minus_white_mean_log2fpkm'])>0 for r in fls6)}/{len(fls6)}",
      'S6_LAR_positive_gene_stage':f"{sum(float(r['pink_minus_white_mean_log2fpkm'])>0 for r in lar6)}/{len(lar6)}",
      'S6_anthocyanin_minus_flavonol_positive_stages':f"{sum(r['contrast']=='anthocyanin_minus_flavonol' and float(r['difference_of_module_log2_effects'])>0 for r in c6)}/{sum(r['contrast']=='anthocyanin_minus_flavonol' for r in c6)}",
      'S6_anthocyanin_minus_PA_positive_stages':f"{sum(r['contrast']=='anthocyanin_minus_proanthocyanidin' and float(r['difference_of_module_log2_effects'])>0 for r in c6)}/{sum(r['contrast']=='anthocyanin_minus_proanthocyanidin' for r in c6)}",
      'claim_ceiling':'one reported C. sinensis white-pink system; quantitative within-study evidence, not an independent cross-study meta effect'}
    (a.out_dir/'summary.json').write_text(json.dumps(summary,indent=2)+'\n');print(json.dumps(summary,indent=2))
if __name__=='__main__':main()
