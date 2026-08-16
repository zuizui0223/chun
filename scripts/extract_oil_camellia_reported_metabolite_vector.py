#!/usr/bin/env python3
"""Quantify W/P/CP/R oil-Camellia metabolite-state vector from Zeng2024.

The source has an internal taxon-name conflict between Materials and Fig. 1.
Therefore this script admits only the author-defined colour-material labels
W=white, P=pink, CP=candy pink, R=red and makes no species-level inference.
The entire supplement Table S2 is used (3 replicates per colour material).
"""
from __future__ import annotations
import argparse,csv,json,math
from pathlib import Path
from statistics import mean,stdev
from openpyxl import load_workbook
from scipy.stats import spearmanr

GROUPS={'W_white':['W-1','W-2','W-3'],'P_pink':['P-1','P-2','P-3'],'CP_candy_pink':['CP-1','CP-2','CP-3'],'R_red':['R-1','R-2','R-3']}
ORDER=list(GROUPS)
TARGET={'Anthocyanidins':'anthocyanin_dimension','Flavonols':'flavonol_branch','Proanthocyanidins':'proanthocyanidin_branch','Flavanols':'flavanol_branch'}

def fnum(x):
    try:return float(x or 0)
    except Exception:return 0.0

def hedges_g(a,b):
    if len(a)<2 or len(b)<2:return None
    sp=math.sqrt(((len(a)-1)*stdev(a)**2+(len(b)-1)*stdev(b)**2)/(len(a)+len(b)-2))
    if sp==0:return None
    return (1-3/(4*(len(a)+len(b))-9))*(mean(a)-mean(b))/sp

def read(path):
    wb=load_workbook(path,read_only=True,data_only=True);ws=wb.active
    h=[str(x or '').strip() for x in next(ws.iter_rows(min_row=2,max_row=2,values_only=True))];idx={x:i for i,x in enumerate(h)};rows=[]
    for r in ws.iter_rows(min_row=3,values_only=True):
        if not r[0]:continue
        rows.append({x:r[i] for x,i in idx.items() if x})
    return rows

def main():
    p=argparse.ArgumentParser();p.add_argument('--table-s2',type=Path,required=True);p.add_argument('--out-dir',type=Path,required=True);a=p.parse_args();rows=read(a.table_s2)
    effects=[];compound=[]
    for source_class,dimension in TARGET.items():
        rs=[r for r in rows if str(r.get('Class II') or '')==source_class]
        reps={g:[sum(fnum(r.get(s)) for r in rs) for s in samples] for g,samples in GROUPS.items()}
        means={g:mean(v) for g,v in reps.items()};logs={g:[math.log2(x+1) for x in v] for g,v in reps.items()}
        ranks=[i for i,g in enumerate(ORDER) for _ in logs[g]];vals=[x for g in ORDER for x in logs[g]];rho,pv=spearmanr(ranks,vals)
        rw=math.log2((means['R_red']+1)/(means['W_white']+1))
        effects.append({'independence_cluster':'OIL_CAMELLIA_MULTI','taxon_admission':'colour-material labels only; source taxon names conflicted','source_class':source_class,'state_dimension':dimension,'n_compounds':len(rs),**{f'{g}_mean_total_intensity':means[g] for g in ORDER},'R_red_minus_W_white_log2_total':rw,'R_vs_W_hedges_g_log2_total':hedges_g(logs['R_red'],logs['W_white']),'spearman_colour_order_vs_log2_rep_total':rho,'spearman_p':pv,'claim_ceiling':'within-source colour-material state vector; no species-level comparison because taxon labels conflict'})
        for r in rs:
            cm={g:mean([fnum(r.get(s)) for s in samples]) for g,samples in GROUPS.items()}
            compound.append({'independence_cluster':'OIL_CAMELLIA_MULTI','compound':str(r.get('Compounds') or ''),'source_class':source_class,'state_dimension':dimension,**{f'{g}_mean_intensity':cm[g] for g in ORDER},'R_gt_W':cm['R_red']>cm['W_white'],'monotonic_W_P_CP_R':all(cm[ORDER[i]]<=cm[ORDER[i+1]] for i in range(len(ORDER)-1)),'claim_ceiling':'compound-level within-source diagnostic; compounds are not independent studies'})
    look={r['state_dimension']:r for r in effects};anth=look['anthocyanin_dimension'];contr=[]
    for dim,label in [('flavonol_branch','anthocyanin_minus_flavonol'),('proanthocyanidin_branch','anthocyanin_minus_proanthocyanidin')]:
        x=look[dim]
        contr.append({'independence_cluster':'OIL_CAMELLIA_MULTI','contrast':'R_red_minus_W_white','state_vector_axis':label,'difference_of_log2_total_effects':float(anth['R_red_minus_W_white_log2_total'])-float(x['R_red_minus_W_white_log2_total']),'claim_ceiling':'within-source relative allocation effect using colour-material labels only'})
    a.out_dir.mkdir(parents=True,exist_ok=True)
    def write(name,data):
        with (a.out_dir/name).open('w',newline='',encoding='utf-8') as f:w=csv.DictWriter(f,fieldnames=list(data[0]));w.writeheader();w.writerows(data)
    write('metabolite_class_effects.csv',effects);write('compound_effects.csv',compound);write('state_vector_contrasts.csv',contr)
    fl=look['flavonol_branch'];pa=look['proanthocyanidin_branch']
    summary={'independence_cluster':'OIL_CAMELLIA_MULTI','phenotype_order':'W=white; P=pink; CP=candy pink; R=red','taxon_provenance_status':'conflicted in source; taxon names not admitted for this analysis','n_metabolites':len(rows),'anthocyanin_compounds':anth['n_compounds'],'flavonol_compounds':fl['n_compounds'],'PA_compounds':pa['n_compounds'],'anthocyanin_R_W_log2_total':anth['R_red_minus_W_white_log2_total'],'flavonol_R_W_log2_total':fl['R_red_minus_W_white_log2_total'],'PA_R_W_log2_total':pa['R_red_minus_W_white_log2_total'],'anthocyanin_colour_order_rho':anth['spearman_colour_order_vs_log2_rep_total'],'flavonol_colour_order_rho':fl['spearman_colour_order_vs_log2_rep_total'],'anthocyanin_minus_flavonol_R_W':next(r['difference_of_log2_total_effects'] for r in contr if r['state_vector_axis']=='anthocyanin_minus_flavonol'),'anthocyanin_minus_PA_R_W':next(r['difference_of_log2_total_effects'] for r in contr if r['state_vector_axis']=='anthocyanin_minus_proanthocyanidin'),'claim_ceiling':'one OIL_CAMELLIA_MULTI colour-material cluster; no species-level effect because source taxon naming is inconsistent'}
    (a.out_dir/'summary.json').write_text(json.dumps(summary,indent=2)+'\n');print(json.dumps(summary,indent=2))
if __name__=='__main__':main()
