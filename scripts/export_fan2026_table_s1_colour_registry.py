#!/usr/bin/env python3
"""Export the complete same-workbook Fan2026 Table S1 Camellia colour registry.

This is an audit aid only. It does not infer aliases between Table S1 and S2.
"""
from __future__ import annotations
import argparse,csv,re
from pathlib import Path
from openpyxl import load_workbook

def clean(x): return re.sub(r"\s+"," ",str(x or "").replace("\xa0"," ").strip())

def main():
    p=argparse.ArgumentParser();p.add_argument('--workbook',type=Path,required=True);p.add_argument('--output',type=Path,required=True);a=p.parse_args()
    wb=load_workbook(a.workbook,read_only=True,data_only=True);ws=wb['Table.S1'];rows=list(ws.iter_rows(values_only=True));hi=None;header=None
    for i,r in enumerate(rows):
        vals=[clean(x) for x in r]
        if {'Species','Petal color'}.issubset(set(vals)):hi=i;header=vals;break
    if hi is None:raise SystemExit('Table S1 header not found')
    idx={h:j for j,h in enumerate(header) if h};out=[]
    for r in rows[hi+1:]:
        vals=[clean(x) for x in r]
        if not any(vals):continue
        if 'Genus' in idx and vals[idx['Genus']]!='Camellia':continue
        name=vals[idx['Species']];colour=vals[idx['Petal color']]
        if not name or not colour:continue
        out.append({'source_species_label':name,'petal_colour':colour,'section':vals[idx['Section']] if 'Section' in idx else '','region':vals[idx['Region']] if 'Region' in idx else ''})
    a.output.parent.mkdir(parents=True,exist_ok=True)
    with a.output.open('w',newline='',encoding='utf-8') as f:
        w=csv.DictWriter(f,fieldnames=list(out[0]));w.writeheader();w.writerows(out)
    print(f'exported {len(out)} Table S1 Camellia colour rows')
if __name__=='__main__':main()
