#!/usr/bin/env python3
"""Build a conservative species-level Camellia colour seed from Fan et al. 2026 Data S1.

Source workbook fields are preserved conceptually; this script only normalizes
sample suffixes and genus abbreviations, then collapses infraspecific/sample rows
to the species binomial when all observed source rows agree on the visible colour
state. Species with conflicting states are excluded from the deterministic seed.
"""
from __future__ import annotations
import argparse, json, pathlib, re
from collections import Counter, defaultdict
from openpyxl import load_workbook

COUNTRY = {"China":"CN", "Vietnam":"VN", "Japan":"JP", "China(Taiwan)":"TW"}

def clean(v):
    return "" if v is None else re.sub(r"\s+"," ",str(v).replace("\xa0"," ").strip())

def taxon_unit(s: str) -> str:
    s=clean(s)
    s=re.sub(r"[_-]\d+$","",s)
    if s.startswith("C. "): s="Camellia "+s[3:]
    elif s.startswith("C."): s="Camellia "+s[2:].lstrip()
    return s

def species_base(s: str) -> str | None:
    if not s or "/" in s: return None
    # Remove a trailing quoted cultivar before collapsing infraspecific ranks.
    s=re.sub(r"'[^']+'$","",s).strip()
    s=re.split(r"\s+(?:var\.|subsp\.|ssp\.|f\.|forma|from\.)\s+",s,maxsplit=1,flags=re.I)[0]
    p=s.split()
    return " ".join(p[:2]) if len(p)>=2 and p[0]=="Camellia" else None

def state(x: str) -> str | None:
    x=clean(x).lower()
    if x=="white": return "W"
    if x=="yellow": return "Y"
    if x in {"red","purple red","pink","plink"}: return "A"
    return None

def main():
    ap=argparse.ArgumentParser(); ap.add_argument("xlsx",type=pathlib.Path); ap.add_argument("--output",type=pathlib.Path,required=True); ap.add_argument("--audit",type=pathlib.Path,required=True); a=ap.parse_args()
    wb=load_workbook(a.xlsx,read_only=True,data_only=True); ws=wb["Table.S1"]
    rows=list(ws.iter_rows(values_only=True))
    # Data S1 uses row 2 as header; locate it by the required column labels.
    hi=None; header=None
    for i,r in enumerate(rows):
        vals=[clean(x) for x in r]
        if {"Genus","Species","Petal color","Country"}.issubset(set(vals)):
            hi=i; header=vals; break
    if hi is None: raise SystemExit("Table.S1 required header not found")
    idx={h:j for j,h in enumerate(header) if h}
    by_species=defaultdict(lambda:{"states":set(),"source_colors":set(),"countries":set(),"sections":set(),"areas":set(),"units":set(),"n":0})
    exclusions=[]
    n_camellia=0
    for r in rows[hi+1:]:
        vals=[clean(x) for x in r]
        if not any(vals): continue
        if vals[idx["Genus"]] != "Camellia": continue
        n_camellia += 1
        unit=taxon_unit(vals[idx["Species"]]); sp=species_base(unit); st=state(vals[idx["Petal color"]])
        if not sp or not st:
            exclusions.append({"source_taxon":unit,"reason":"ambiguous_or_unmapped_taxon_or_colour","source_colour":vals[idx["Petal color"]]}); continue
        d=by_species[sp]; d["states"].add(st); d["source_colors"].add(vals[idx["Petal color"]].lower()); d["units"].add(unit); d["n"]+=1
        c=vals[idx["Country"]]
        if c: d["countries"].add(COUNTRY.get(c,c))
        if "Section" in idx and vals[idx["Section"]]: d["sections"].add(vals[idx["Section"]])
        if "Area" in idx and vals[idx["Area"]]: d["areas"].add(vals[idx["Area"]])
    out=[]
    polymorphic=[]
    for sp,d in sorted(by_species.items()):
        if len(d["states"]) != 1:
            polymorphic.append({"taxon":sp,"states":";".join(sorted(d["states"])),"source_colours":";".join(sorted(d["source_colors"])),"source_units":";".join(sorted(d["units"]))}); continue
        st=next(iter(d["states"]))
        out.append({
            "taxon":sp,
            "colour_state":st,
            "pigment_proxy":{"A":"visible_anthocyanin_like","W":"visible_white","Y":"visible_yellow"}[st],
            "native_country_codes":";".join(sorted(d["countries"])),
            "section":";".join(sorted(d["sections"],key=str.lower)),
            "areas":";".join(sorted(d["areas"])),
            "n_source_accessions":d["n"],
            "source_color_values":";".join(sorted(d["source_colors"])),
            "source_units":";".join(sorted(d["units"])),
            "source_basis":"Fan2026 Data S1 taxon-level petal-colour evidence",
            "source_doi_or_registry":"10.1111/pbi.70442;PBI-24-1725-s002.xlsx",
            "analysis_role":"fan2026_species_screen",
            "notes":"country filters are source-provenance countries, not asserted complete native ranges",
        })
    a.output.parent.mkdir(parents=True,exist_ok=True)
    import csv
    fields=list(out[0])
    with a.output.open("w",newline="",encoding="utf-8") as f:
        w=csv.DictWriter(f,fieldnames=fields); w.writeheader(); w.writerows(out)
    summary={
        "source_camellia_accession_rows":n_camellia,
        "n_species_with_records":len(by_species),
        "n_stable_species_admitted":len(out),
        "states":dict(Counter(r["colour_state"] for r in out)),
        "n_polymorphic_species_excluded":len(polymorphic),
        "polymorphic_species":polymorphic,
        "other_exclusions":exclusions,
        "normalization":"sample suffixes _N/-N removed; genus C. expanded; infraspecific/cultivar rows collapsed only when species-level source states agree",
    }
    a.audit.parent.mkdir(parents=True,exist_ok=True); a.audit.write_text(json.dumps(summary,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    print(json.dumps(summary,ensure_ascii=False,indent=2))
if __name__=="__main__": main()
