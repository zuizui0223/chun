#!/usr/bin/env python3
"""Quantify Zhou2020 white(ZJW)-pink(BTP) pathway expression with provenance gates.

Table S6 has a systematic BTP/ZJW header reversal relative to Table S8, Table S9,
and the article text. S6 is admitted only after exact overlapping triplets prove the
swap. Flavonol (FLS) and proanthocyanidin (LAR/ANR) branches are kept separate.
All rows belong to one independence cluster; genes/stages are repeated measures.
"""
from __future__ import annotations
import argparse,csv,json,math,re
from collections import defaultdict
from pathlib import Path
from statistics import mean,median,stdev
from openpyxl import load_workbook

STAGES=range(1,6)
# LAR must precede ANR: 'leucoanthocyanidin reductase' contains the substring
# 'anthocyanidin reductase' and otherwise can be misclassified as ANR.
PATTERNS=[
 ('DFR',re.compile(r'\bDFR\b|DIHYDROFLAVONOL[- ]?4[- ]?REDUCTASE',re.I)),
 ('ANS',re.compile(r'\bANS\b|\bANTHOCYANIDIN SYNTHASE\b',re.I)),
 ('UFGT',re.compile(r'\bUFGT\b|UDP.*GLUCOSYLTRANSFERASE',re.I)),
 ('FLS',re.compile(r'\bFLS\d*\b|FLAVONOL SYNTHASE',re.I)),
 ('LAR',re.compile(r'\bF?LAR\b|LEUCOANTHOCYANIDIN REDUCT',re.I)),
 ('ANR',re.compile(r'\bANR\b|\bANTHOCYANIDIN REDUCTASE\b',re.I)),
 ('CHS',re.compile(r'\bCHS\d*\b|CHALCONE SYNTHASE',re.I)),
 ('F3H_5PRIME',re.compile(r"F3['’]5['’]H",re.I)),
 ('F3H_PRIME',re.compile(r"F3['’]H",re.I)),
 ('F3H',re.compile(r'\bF3H\b|FLAVANONE 3[- ]HYDROXYLASE',re.I)),
]

def gclass(x):
    s=str(x or '').strip()
    if not s or s=='--':return None
    for name,p in PATTERNS:
        if p.search(s):return name
    return None

def fnum(x):
    try:
        v=float(x);return v if math.isfinite(v) and v>=0 else None
    except Exception:return None

def close3(a,b,tol=1e-7):
    return len(a)==len(b)==3 and all(abs(x-y)<=tol*max(1.0,abs(x),abs(y)) for x,y in zip(a,b))

def log_effect(white,pink):
    lw=[math.log2(x+1) for x in white];lp=[math.log2(x+1) for x in pink]
    return mean(lp)-mean(lw),lw,lp

def hedges_g(lp,lw):
    if len(lp)<2 or len(lw)<2:return None
    sp2=((len(lp)-1)*stdev(lp)**2+(len(lw)-1)*stdev(lw)**2)/(len(lp)+len(lw)-2)
    if sp2<=0:return None
    return (1-3/(4*(len(lp)+len(lw))-9))*(mean(lp)-mean(lw))/math.sqrt(sp2)

def module(gc):
    if gc in {'DFR','ANS','UFGT'}:return 'anthocyanin_downstream'
    if gc=='FLS':return 'flavonol_branch'
    if gc in {'LAR','ANR'}:return 'proanthocyanidin_branch'
    if gc in {'CHS','F3H','F3H_PRIME','F3H_5PRIME'}:return 'shared_or_upstream_flavonoid'
    return None

def effect_row(gid,gc,stage,source,white,pink,reported_log2fc='',fdr=''):
    d,lw,lp=log_effect(white,pink)
    return {'independence_cluster':'CSIN_WHITE_PINK','gene_id':gid,'gene_class':gc,'stage':stage,'source_table':source,
      'mean_white_fpkm':mean(white),'mean_pink_fpkm':mean(pink),'pink_minus_white_mean_log2fpkm':d,
      'hedges_g_log2fpkm':hedges_g(lp,lw),'author_reported_log2fc':reported_log2fc,'author_reported_fdr':fdr,
      'claim_ceiling':'one within-study processed-expression contrast; genes/stages are repeated measures'}

def load_s8(path):
    wb=load_workbook(path,read_only=True,data_only=True);ann={};prof={};rows=[]
    for st in STAGES:
        ws=next(x for x in wb.worksheets if f'stage{st}' in x.title.lower())
        for r in ws.iter_rows(values_only=True):
            gid=str(r[0] or '').strip()
            if not gid or gid.startswith('#'):continue
            gc=gclass(r[1] if len(r)>1 else None)
            if not gc or len(r)<14:continue
            white=[fnum(r[i]) for i in (2,4,6)];pink=[fnum(r[i]) for i in (8,10,12)]
            if any(x is None for x in white+pink):continue
            ann[gid]=gc;prof[(gid,st)]=(white,pink)
            lfc=float(r[15]) if len(r)>15 and r[15] is not None else ''
            fdr=fnum(r[14]) if len(r)>14 else ''
            rows.append(effect_row(gid,gc,st,'S8_stage_selected_DEG',white,pink,lfc,fdr))
    return ann,prof,rows

def find_header(ws,pred):
    for i,r in enumerate(ws.iter_rows(min_row=1,max_row=8,values_only=True),1):
        vals=[str(x or '').strip() for x in r]
        if pred(vals):return i,vals
    raise SystemExit('header not found')

def load_s6_raw(path):
    wb=load_workbook(path,read_only=True,data_only=True);ws=wb.active
    hr,h=find_header(ws,lambda v:'#ID' in v and 'BTP11_FPKM' in v and 'ZJW11_FPKM' in v);idx={x:i for i,x in enumerate(h)}
    ann={};data={}
    for r in ws.iter_rows(min_row=hr+1,values_only=True):
        gid=str(r[idx['#ID']] or '').strip();gc=gclass(r[idx['gene_name']] if gid else None)
        if not gid or not gc:continue
        ann[gid]=gc
        for st in STAGES:
            lab_b=[fnum(r[idx[f'BTP{st}{q}_FPKM']]) for q in (1,2,3)]
            lab_z=[fnum(r[idx[f'ZJW{st}{q}_FPKM']]) for q in (1,2,3)]
            if not any(x is None for x in lab_b+lab_z):data[(gid,st)]=(lab_z,lab_b)
    return ann,data

def audit_s6(data,s8):
    direct=swapped=0;examples=[]
    for key,(lab_z,lab_b) in data.items():
        if key not in s8:continue
        w,p=s8[key];d=close3(lab_z,w) and close3(lab_b,p);s=close3(lab_b,w) and close3(lab_z,p)
        direct+=int(d);swapped+=int(s);examples.append({'gene_id':key[0],'stage':key[1],'direct_match':d,'swapped_match':s})
    if swapped<3 or swapped<=direct:raise SystemExit(f'S6 label swap gate failed direct={direct} swapped={swapped}')
    return {'table':'S6','overlap_tests':len(examples),'direct_matches':direct,'swapped_matches':swapped,
      'decision':'S6 headers are systematically reversed: printed BTP=biological ZJW/white; printed ZJW=biological BTP/pink'},examples

def s6_effects(data,ann):
    out=[]
    for (gid,st),(lab_z,lab_b) in sorted(data.items(),key=lambda x:(x[0][1],x[0][0])):
        out.append(effect_row(gid,ann[gid],st,'S6_hub_full_profile_label_swap_corrected',lab_b,lab_z))
    return out

def load_s9(path,ann,s8):
    wb=load_workbook(path,read_only=True,data_only=True);ws=wb.active
    hr,h=find_header(ws,lambda v:'modColor' in v and 'ZJW11' in v and 'BTP11' in v);idx={x:i for i,x in enumerate(h)}
    rows=[];direct=swapped=0
    for r in ws.iter_rows(min_row=hr+1,values_only=True):
        gid=str(r[0] or '').strip()
        if gid not in ann:continue
        for st in STAGES:
            kz=[f'ZJW{st}{q}' for q in (1,2,3)];kb=[f'BTP{st}{q}' for q in (1,2,3)]
            if not all(k in idx for k in kz+kb):continue
            w=[fnum(r[idx[k]]) for k in kz];p=[fnum(r[idx[k]]) for k in kb]
            if any(x is None for x in w+p):continue
            if (gid,st) in s8:
                sw,sp=s8[(gid,st)];direct+=int(close3(w,sw) and close3(p,sp));swapped+=int(close3(p,sw) and close3(w,sp))
            rows.append(effect_row(gid,ann[gid],st,'S9_module_full_profile',w,p))
    return rows,{'table':'S9','overlap_tests':direct+swapped,'direct_matches':direct,'swapped_matches':swapped,'decision':'printed labels retained'}

def module_summary(rows,source):
    by=defaultdict(list);out=[]
    for r in rows:
        m=module(r['gene_class'])
        if m:by[(r['stage'],m)].append(r)
    for (st,m),rs in sorted(by.items()):
        v=[float(r['pink_minus_white_mean_log2fpkm']) for r in rs]
        out.append({'independence_cluster':'CSIN_WHITE_PINK','source_set':source,'stage':st,'module':m,'n_genes':len(v),
          'mean_gene_log2_effect':mean(v),'median_gene_log2_effect':median(v),'n_positive_genes':sum(x>0 for x in v),'n_negative_genes':sum(x<0 for x in v),
          'gene_ids':';'.join(r['gene_id'] for r in rs),'claim_ceiling':'descriptive within-study selected-gene module summary'})
    return out

def module_contrasts(ms):
    L={(r['source_set'],r['stage'],r['module']):r for r in ms};out=[]
    for src in sorted({r['source_set'] for r in ms}):
      for st in STAGES:
        a=L.get((src,st,'anthocyanin_downstream'))
        for cname,m in [('anthocyanin_minus_flavonol','flavonol_branch'),('anthocyanin_minus_proanthocyanidin','proanthocyanidin_branch')]:
            x=L.get((src,st,m))
            if a and x:out.append({'independence_cluster':'CSIN_WHITE_PINK','source_set':src,'stage':st,'contrast':cname,
              'difference_of_module_log2_effects':float(a['mean_gene_log2_effect'])-float(x['mean_gene_log2_effect']),
              'claim_ceiling':'relative within-study allocation contrast; FLS and PA kept separate'})
    return out

def write(path,rows):
    path.parent.mkdir(parents=True,exist_ok=True);fields=list(rows[0]) if rows else ['empty']
    with path.open('w',newline='',encoding='utf-8') as f:
        w=csv.DictWriter(f,fieldnames=fields);w.writeheader();w.writerows(rows)

def main():
    p=argparse.ArgumentParser();p.add_argument('--s6',type=Path,required=True);p.add_argument('--s8',type=Path,required=True);p.add_argument('--s9',type=Path,required=True);p.add_argument('--out-dir',type=Path,required=True);a=p.parse_args()
    ann8,p8,r8=load_s8(a.s8);ann6,d6=load_s6_raw(a.s6);ann={**ann8,**ann6};gate6,examples=audit_s6(d6,p8);r6=s6_effects(d6,ann6);r9,gate9=load_s9(a.s9,ann,p8)
    m6=module_summary(r6,'S6_hub_full_profile_label_swap_corrected');m8=module_summary(r8,'S8_stage_selected_DEG');m9=module_summary(r9,'S9_module_full_profile');cs=module_contrasts(m6+m8+m9)
    write(a.out_dir/'s6_hub_gene_stage_effects.csv',r6);write(a.out_dir/'s8_selected_deg_effects.csv',r8);write(a.out_dir/'s9_module_gene_stage_effects.csv',r9);write(a.out_dir/'module_stage_effects.csv',m6+m8+m9);write(a.out_dir/'module_contrasts.csv',cs);write(a.out_dir/'s6_label_swap_examples.csv',examples)
    (a.out_dir/'provenance_label_audit.json').write_text(json.dumps({'S6':gate6,'S9':gate9},indent=2)+'\n')
    def pos(rows,g):
        x=[r for r in rows if r['gene_class']==g];return f"{sum(float(r['pink_minus_white_mean_log2fpkm'])>0 for r in x)}/{len(x)}"
    c6=[r for r in cs if r['source_set']=='S6_hub_full_profile_label_swap_corrected']
    def pc(name):
        x=[r for r in c6 if r['contrast']==name];return f"{sum(float(r['difference_of_module_log2_effects'])>0 for r in x)}/{len(x)}"
    summary={'independence_cluster':'CSIN_WHITE_PINK','s6_label_gate':gate6,'s9_label_gate':gate9,'S6_DFR_positive_gene_stage':pos(r6,'DFR'),'S6_FLS_positive_gene_stage':pos(r6,'FLS'),'S6_LAR_positive_gene_stage':pos(r6,'LAR'),'S6_ANR_positive_gene_stage':pos(r6,'ANR'),'S6_anthocyanin_minus_flavonol_positive_stages':pc('anthocyanin_minus_flavonol'),'S6_anthocyanin_minus_PA_positive_stages':pc('anthocyanin_minus_proanthocyanidin'),'claim_ceiling':'one C. sinensis system; quantitative within-study evidence, not cross-study pooled effect'}
    (a.out_dir/'summary.json').write_text(json.dumps(summary,indent=2)+'\n');print(json.dumps(summary,indent=2))
if __name__=='__main__':main()
