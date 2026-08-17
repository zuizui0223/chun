#!/usr/bin/env python3
"""Scan public OA supplement packages for exact biological terms and ID-like context.

Supports text/CSV/TSV, XLSX and PDF. This is a provenance discovery tool for
source-reported locus/transcript identifiers. It does not OCR images and does
not infer missing IDs.
"""
from __future__ import annotations
import argparse,csv,json,re
from pathlib import Path
from openpyxl import load_workbook
from pypdf import PdfReader

def norm(s):return re.sub(r'\s+',' ',str(s or '')).strip()
def contexts(text,terms,window=500):
    t=norm(text);low=t.lower();out=[]
    for term in terms:
        pos=0;q=term.lower()
        while True:
            i=low.find(q,pos)
            if i<0:break
            out.append({'term':term,'context':t[max(0,i-window):min(len(t),i+len(term)+window)]});pos=i+len(q)
    return out

def main():
    ap=argparse.ArgumentParser();ap.add_argument('--root',type=Path,required=True);ap.add_argument('--terms',nargs='+',required=True);ap.add_argument('--output',type=Path,required=True);a=ap.parse_args()
    hits=[];fail=[]
    for path in sorted(p for p in a.root.rglob('*') if p.is_file()):
        suf=path.suffix.lower();rel=str(path.relative_to(a.root))
        try:
            if suf in {'.txt','.csv','.tsv','.xml','.html','.htm'}:
                for h in contexts(path.read_text(encoding='utf-8',errors='replace'),a.terms):hits.append({'file':rel,'kind':'text',**h})
            elif suf=='.xlsx':
                wb=load_workbook(path,read_only=True,data_only=True)
                for ws in wb.worksheets:
                    for rno,row in enumerate(ws.iter_rows(values_only=True),1):
                        text=' | '.join(norm(x) for x in row if x is not None)
                        for h in contexts(text,a.terms,350):hits.append({'file':rel,'kind':'xlsx','sheet':ws.title,'row':rno,**h})
            elif suf=='.pdf':
                rd=PdfReader(str(path))
                for pno,page in enumerate(rd.pages,1):
                    for h in contexts(page.extract_text() or '',a.terms,600):hits.append({'file':rel,'kind':'pdf','page':pno,**h})
        except Exception as e:fail.append({'file':rel,'error':repr(e)})
    result={'terms':a.terms,'hit_count':len(hits),'hits':hits,'failures':fail,'claim_ceiling':'exact term/source-ID discovery only; no OCR and no inferred sequence identity'}
    a.output.parent.mkdir(parents=True,exist_ok=True);a.output.write_text(json.dumps(result,indent=2,ensure_ascii=False)+'\n',encoding='utf-8');print(json.dumps({'hit_count':len(hits),'failure_count':len(fail)},indent=2))
if __name__=='__main__':main()
