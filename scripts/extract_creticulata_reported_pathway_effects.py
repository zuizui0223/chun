#!/usr/bin/env python3
"""Quantify the independent C. reticulata red/pink/white pigment-state system.

Source: Geng et al. 2022, PMC9725097. HHYC=red, XJ=pink, TZM=white.

The script deliberately separates three evidence layers:
1. four late anthocyanin UFGT loci highlighted by the source study;
2. broad KEGG-annotated pigment-gene families as a paralog-heterogeneity audit;
3. measured metabolite-class totals (Anthocyanins, Flavonols,
   Proanthocyanidins, Flavanols).

This prevents broad homolog families from being treated as interchangeable
functional orthologs. All outputs belong to one independence cluster,
CRETICULATA.
"""
from __future__ import annotations

import argparse,csv,json,math,re
from collections import defaultdict
from pathlib import Path
from statistics import mean,median,stdev

from openpyxl import load_workbook

SAMPLES=["HHYC1","HHYC2","HHYC3","XJ1","XJ2","XJ3","TZM1","TZM2","TZM3"]
GROUPS={"red_HHYC":slice(0,3),"pink_XJ":slice(3,6),"white_TZM":slice(6,9)}
KEY_UFGT={"gene-LOC114289234","gene-LOC114285765","gene-LOC114285774","gene-LOC114288492"}
KMAP={
 "K13082":("DFR_like","anthocyanin_downstream"),
 "K05277":("ANS_like","anthocyanin_downstream"),
 "K12930":("UFGT_like","anthocyanin_downstream"),
 "K05278":("FLS_like","flavonol_branch"),
 "K13081":("LAR_like","proanthocyanidin_branch"),
 "K08695":("ANR_like","proanthocyanidin_branch"),
 "K00660":("CHS_like","shared_upstream"),
 "K01859":("CHI_like","shared_upstream"),
 "K00475":("F3H_like","shared_upstream"),
}

def safe_float(x):
    try:return float(x or 0)
    except Exception:return 0.0

def hedges_g(a,b):
    a=list(a);b=list(b)
    if len(a)<2 or len(b)<2:return None
    va=stdev(a)**2;vb=stdev(b)**2;df=len(a)+len(b)-2
    sp=math.sqrt(((len(a)-1)*va+(len(b)-1)*vb)/df)
    if sp==0:return None
    d=(mean(a)-mean(b))/sp;j=1-3/(4*(len(a)+len(b))-9)
    return j*d

def read_table1(path):
    wb=load_workbook(path,read_only=True,data_only=True);ws=wb.active
    rows=[]
    for r in ws.iter_rows(min_row=4,values_only=True):
        gid=str(r[0] or '').strip()
        if not gid:continue
        vals=[safe_float(x) for x in r[1:10]]
        kegg=str(r[24] or '')
        codes=re.findall(r'\bK\d{5}\b',kegg)
        rows.append({'gene_id':gid,'values':vals,'kegg_codes':codes,'gene_name':str(r[23] or ''),'kegg_annotation':kegg,'swissprot':str(r[27] or '')})
    return rows

def zscore(values):
    x=[math.log2(v+1) for v in values];m=mean(x);sd=stdev(x)
    return [(v-m)/sd for v in x] if sd>0 else [0.0]*len(x)

def key_ufgt_summary(rows):
    hit=[r for r in rows if r['gene_id'] in KEY_UFGT]
    missing=sorted(KEY_UFGT-{r['gene_id'] for r in hit})
    if missing:raise SystemExit(f'missing source-highlighted UFGT loci: {missing}')
    gene_rows=[];zrows=[]
    for r in sorted(hit,key=lambda x:x['gene_id']):
        v=r['values'];lg=[math.log2(x+1) for x in v];z=zscore(v);zrows.append(z)
        gm={g:mean(lg[s]) for g,s in GROUPS.items()}
        gene_rows.append({'independence_cluster':'CRETICULATA','gene_id':r['gene_id'],'node':'late_UFGT','red_mean_log2fpkm':gm['red_HHYC'],'pink_mean_log2fpkm':gm['pink_XJ'],'white_mean_log2fpkm':gm['white_TZM'],'red_minus_white_log2fpkm':gm['red_HHYC']-gm['white_TZM'],'red_gt_white':gm['red_HHYC']>gm['white_TZM'],'red_gt_pink_gt_white':gm['red_HHYC']>gm['pink_XJ']>gm['white_TZM'],'claim_ceiling':'source-highlighted late UFGT loci within one C. reticulata system'})
    module=[mean(col) for col in zip(*zrows)]
    groups={g:module[s] for g,s in GROUPS.items()}
    contrasts=[]
    for a,b,name in [('red_HHYC','white_TZM','red_minus_white'),('pink_XJ','white_TZM','pink_minus_white'),('red_HHYC','pink_XJ','red_minus_pink')]:
        contrasts.append({'independence_cluster':'CRETICULATA','module':'source_highlighted_late_UFGT','contrast':name,'group_a_mean_z':mean(groups[a]),'group_b_mean_z':mean(groups[b]),'difference_z':mean(groups[a])-mean(groups[b]),'hedges_g':hedges_g(groups[a],groups[b]),'n_genes':len(hit),'n_replicates_per_group':3,'claim_ceiling':'within-study replicate-standardized four-locus module effect; not a cross-study pooled effect'})
    return gene_rows,contrasts

def broad_homolog_audit(rows,min_max_mean=1.0):
    out=[]
    for r in rows:
        codes=[c for c in r['kegg_codes'] if c in KMAP]
        if not codes:continue
        v=r['values'];means=[mean(v[0:3]),mean(v[3:6]),mean(v[6:9])]
        if max(means)<min_max_mean:continue
        lg=[math.log2(x+1) for x in v];lm=[mean(lg[0:3]),mean(lg[3:6]),mean(lg[6:9])]
        for c in codes:
            gc,mod=KMAP[c]
            out.append({'independence_cluster':'CRETICULATA','gene_id':r['gene_id'],'kegg_code':c,'gene_class':gc,'module':mod,'max_group_mean_fpkm':max(means),'red_minus_white_log2fpkm':lm[0]-lm[2],'red_gt_white':lm[0]>lm[2],'red_gt_pink_gt_white':lm[0]>lm[1]>lm[2],'claim_ceiling':'KEGG-annotated homolog-family diagnostic; not functional-ortholog validation'})
    summary=[]
    by=defaultdict(list)
    for r in out:by[(r['module'],r['gene_class'])].append(r)
    for (mod,gc),rs in sorted(by.items()):
        vals=[float(r['red_minus_white_log2fpkm']) for r in rs]
        summary.append({'independence_cluster':'CRETICULATA','module':mod,'gene_class':gc,'n_active_homologs':len(rs),'n_red_gt_white':sum(str(r['red_gt_white']).lower()=='true' if isinstance(r['red_gt_white'],str) else r['red_gt_white'] for r in rs),'n_monotonic_red_pink_white':sum(str(r['red_gt_pink_gt_white']).lower()=='true' if isinstance(r['red_gt_pink_gt_white'],str) else r['red_gt_pink_gt_white'] for r in rs),'median_red_minus_white_log2fpkm':median(vals),'claim_ceiling':'paralog heterogeneity diagnostic; broad KEGG family is not an interchangeable pathway node'})
    return out,summary

def read_metabolites(path):
    wb=load_workbook(path,read_only=True,data_only=True);ws=wb.active;rows=[]
    for r in ws.iter_rows(min_row=3,values_only=True):
        if r[0] is None:continue
        rows.append({'compound_id':str(r[0]),'compound':str(r[1]),'class_i':str(r[2]),'class_ii':str(r[3]),'red_HHYC':safe_float(r[5]),'pink_XJ':safe_float(r[6]),'white_TZM':safe_float(r[7])})
    return rows

def metabolite_summary(rows):
    target=['Anthocyanins','Flavonols','Proanthocyanidins','Flavanols'];out=[];byclass={}
    for cls in target:
        rs=[r for r in rows if r['class_ii']==cls];byclass[cls]=rs
        sums={g:sum(r[g] for r in rs) for g in ['red_HHYC','pink_XJ','white_TZM']}
        out.append({'independence_cluster':'CRETICULATA','metabolite_class':cls,'n_compounds':len(rs),**{f'{g}_total_intensity':v for g,v in sums.items()},'red_minus_white_log2_total':math.log2((sums['red_HHYC']+1)/(sums['white_TZM']+1)),'pink_minus_white_log2_total':math.log2((sums['pink_XJ']+1)/(sums['white_TZM']+1)),'red_minus_pink_log2_total':math.log2((sums['red_HHYC']+1)/(sums['pink_XJ']+1)),'n_red_gt_white':sum(r['red_HHYC']>r['white_TZM'] for r in rs),'n_red_gt_pink_gt_white':sum(r['red_HHYC']>r['pink_XJ']>r['white_TZM'] for r in rs),'claim_ceiling':'within-study targeted metabolite-class totals from reported mean ion intensities; no between-study unit equivalence assumed'})
    lookup={r['metabolite_class']:r for r in out};contr=[]
    for b,label in [('Flavonols','anthocyanin_minus_flavonol'),('Proanthocyanidins','anthocyanin_minus_proanthocyanidin')]:
        a=lookup['Anthocyanins'];x=lookup[b]
        for c in ['red_minus_white','pink_minus_white','red_minus_pink']:
            contr.append({'independence_cluster':'CRETICULATA','contrast':c,'state_vector_axis':label,'difference_of_log2_total_effects':float(a[f'{c}_log2_total'])-float(x[f'{c}_log2_total']),'claim_ceiling':'within-study relative pigment-class allocation effect; classes retained as separate latent-state dimensions'})
    return out,contr

def write(path,rows):
    path.parent.mkdir(parents=True,exist_ok=True);fields=list(rows[0]) if rows else ['empty']
    with path.open('w',newline='',encoding='utf-8') as f:
        w=csv.DictWriter(f,fieldnames=fields);w.writeheader();w.writerows(rows)

def main():
    p=argparse.ArgumentParser();p.add_argument('--table1',type=Path,required=True);p.add_argument('--table4',type=Path,required=True);p.add_argument('--out-dir',type=Path,required=True);a=p.parse_args()
    genes=read_table1(a.table1);keygenes,keyeffects=key_ufgt_summary(genes);broad,broadsummary=broad_homolog_audit(genes);mets=read_metabolites(a.table4);ms,mc=metabolite_summary(mets)
    write(a.out_dir/'key_ufgt_gene_effects.csv',keygenes);write(a.out_dir/'key_ufgt_module_effects.csv',keyeffects);write(a.out_dir/'broad_homolog_audit.csv',broad);write(a.out_dir/'broad_homolog_summary.csv',broadsummary);write(a.out_dir/'metabolite_class_effects.csv',ms);write(a.out_dir/'metabolite_state_vector_contrasts.csv',mc)
    au=next(r for r in ms if r['metabolite_class']=='Anthocyanins');fl=next(r for r in ms if r['metabolite_class']=='Flavonols');pa=next(r for r in ms if r['metabolite_class']=='Proanthocyanidins')
    summary={'independence_cluster':'CRETICULATA','phenotype_mapping':'HHYC=red; XJ=pink; TZM=white','key_ufgt_red_gt_white':f"{sum(r['red_gt_white'] for r in keygenes)}/{len(keygenes)}",'key_ufgt_red_gt_pink_gt_white':f"{sum(r['red_gt_pink_gt_white'] for r in keygenes)}/{len(keygenes)}",'key_ufgt_red_white_hedges_g':next(r['hedges_g'] for r in keyeffects if r['contrast']=='red_minus_white'),'anthocyanin_compounds':au['n_compounds'],'anthocyanin_red_white_log2_total':au['red_minus_white_log2_total'],'anthocyanin_pink_white_log2_total':au['pink_minus_white_log2_total'],'flavonol_red_white_log2_total':fl['red_minus_white_log2_total'],'PA_red_white_log2_total':pa['red_minus_white_log2_total'],'anthocyanin_minus_flavonol_red_white':next(r['difference_of_log2_total_effects'] for r in mc if r['contrast']=='red_minus_white' and r['state_vector_axis']=='anthocyanin_minus_flavonol'),'anthocyanin_minus_PA_red_white':next(r['difference_of_log2_total_effects'] for r in mc if r['contrast']=='red_minus_white' and r['state_vector_axis']=='anthocyanin_minus_proanthocyanidin'),'claim_ceiling':'one independent C. reticulata system; quantitative corroboration of multidimensional pigment allocation, not cross-study pooled effect'}
    a.out_dir.mkdir(parents=True,exist_ok=True);(a.out_dir/'summary.json').write_text(json.dumps(summary,indent=2)+'\n');print(json.dumps(summary,indent=2))
if __name__=='__main__':main()
