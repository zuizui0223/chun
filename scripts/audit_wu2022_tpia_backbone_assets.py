#!/usr/bin/env python3
"""Audit whether Wu et al. 2022 Camellia data can seed a public nuclear backbone.

This is a provenance/crosswalk gate, not a phylogenetic reconstruction. It joins
public taxon evidence from PRJNA665925 RunInfo and, when available, publisher
Table S1 to the live TPIA assembly and bulk-download catalogs.

Important: ARCHIVE_TPIA_CROSSWALK below is a study-specific archive-to-resource
crosswalk. It is NOT asserted as a general taxonomic synonym list.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path
from urllib.parse import quote

import requests
from openpyxl import load_workbook

# Same Wu 2022 / PRJNA665925 resource represented with orthographic or
# infraspecific-name differences between NCBI RunInfo and TPIA.
ARCHIVE_TPIA_CROSSWALK = {
    "Camellia albosericea": ("Camellia albo-sericea", "same_project_orthographic_variant"),
    "Camellia apolyodonta": ("Camellia apolydonta", "same_project_orthographic_variant"),
    "Camellia kissii": ("Camellia kissi", "same_project_orthographic_variant"),
    "Camellia leyensis": ("Camellia leyeensis", "same_project_orthographic_variant"),
    "Camellia pingguoensis": ("Camellia pinggaoensis", "same_project_orthographic_variant"),
    "Camellia stichoclada": ("Camellia stictoclada", "same_project_orthographic_variant"),
    # NCBI ScientificName includes the infraspecific name; TPIA exposes the
    # infraspecific epithet as the assembly name. This is a resource crosswalk,
    # not a taxonomic synonym assertion.
    "Camellia henryana": ("Camellia trichocarpa", "same_project_infraspecific_epithet_resource_name"),
    "Camellia pyxidiacea": ("Camellia rubituberculata", "same_project_infraspecific_epithet_resource_name"),
}


def clean(x):
    return "" if x is None else re.sub(r"\s+", " ", str(x).replace("\xa0", " ").strip())


def norm_taxon(x: str) -> str:
    x = clean(x)
    x = re.sub(r"^C\.\s*", "Camellia ", x)
    x = re.sub(r"\s+", " ", x)
    m = re.search(r"\bCamellia\s+([A-Za-z][A-Za-z-]+)", x, re.I)
    return f"Camellia {m.group(1).lower()}" if m else ""


def display_taxon(x: str) -> str:
    t = norm_taxon(x)
    if not t: return ""
    g, e = t.split(" ", 1)
    return f"{g} {e}"


def extract_table_s1_taxa(xlsx: Path):
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    candidates = []
    for ws in wb.worksheets:
        score = 0
        lname = ws.title.lower().replace(" ", "")
        if "s1" in lname or "table1" in lname: score += 5
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 20: break
            vals = [clean(v) for v in row]
            joined = " ".join(vals).lower()
            if "species" in joined: score += 2
            if "accession" in joined or "bioproject" in joined or "sra" in joined: score += 1
        candidates.append((score, ws))
    candidates.sort(key=lambda z: z[0], reverse=True)
    ws = candidates[0][1]
    rows = list(ws.iter_rows(values_only=True))
    header_i = None
    for i, row in enumerate(rows[:40]):
        vals = [clean(v) for v in row]
        if any(v.lower() == "species" or "species" in v.lower() for v in vals):
            header_i = i; break
    if header_i is None: header_i = 0
    header = [clean(v) or f"col_{j+1}" for j, v in enumerate(rows[header_i])]
    seen = defaultdict(int); unique_header = []
    for h in header:
        seen[h] += 1
        unique_header.append(h if seen[h] == 1 else f"{h}_{seen[h]}")
    out_rows = []; taxa = set()
    for row in rows[header_i + 1:]:
        vals = [clean(v) for v in row]
        if not any(vals): continue
        vals += [""] * (len(unique_header) - len(vals))
        rec = dict(zip(unique_header, vals[:len(unique_header)]))
        found = sorted({t for v in vals if (t := display_taxon(v))})
        rec["_normalized_camellia_taxa"] = ";".join(found)
        taxa.update(found); out_rows.append(rec)
    return ws.title, unique_header + ["_normalized_camellia_taxa"], out_rows, sorted(taxa)


def extract_sra_taxa(path: Path):
    with path.open(newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
    taxa = set(); out = []
    for r in rows:
        sci = clean(r.get("ScientificName") or r.get("scientific_name") or r.get("Organism"))
        tax = display_taxon(sci)
        if tax: taxa.add(tax)
        out.append({
            "Run": clean(r.get("Run")),
            "BioSample": clean(r.get("BioSample")),
            "ScientificName": sci,
            "normalized_taxon": tax,
            "BioProject": clean(r.get("BioProject")),
            "LibraryStrategy": clean(r.get("LibraryStrategy")),
        })
    return out, sorted(taxa)


def fetch_json(session, url, **kwargs):
    r = session.get(url, timeout=90, **kwargs)
    r.raise_for_status()
    try: return r.json()
    except Exception: return json.loads(r.text)


def write_csv(path: Path, rows, fields=None):
    path.parent.mkdir(parents=True, exist_ok=True)
    if fields is None: fields = sorted({k for r in rows for k in r}) if rows else []
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader(); w.writerows(rows)


def bulk_assembly_taxon(description: str, filename: str) -> str:
    text = clean(description)
    m = re.match(r"(?i)^transcriptome assembly of\s+(Camellia\s+.+)$", text)
    if m: return display_taxon(m.group(1))
    # fallback for filename forms like Camellia_foo_Trans.fas.gz
    x = re.sub(r"(?i)(?:_Trans\.fas\.gz|\.zip|\.fas\.gz|\.gz)$", "", clean(filename))
    x = x.replace("_", " ")
    return display_taxon(x)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--supp-xlsx", type=Path)
    ap.add_argument("--sra-runinfo", type=Path)
    ap.add_argument("--out-dir", type=Path, required=True)
    ap.add_argument("--tpia-base", default="https://tpia.teaplants.cn")
    args = ap.parse_args()
    if not args.supp_xlsx and not args.sra_runinfo:
        raise SystemExit("Provide --supp-xlsx and/or --sra-runinfo")
    out = args.out_dir; out.mkdir(parents=True, exist_ok=True)

    source_taxa = set(); source_parts = []
    table_sheet = None; table_rows_n = 0
    if args.supp_xlsx and args.supp_xlsx.exists():
        sheet, fields, rows, taxa = extract_table_s1_taxa(args.supp_xlsx)
        write_csv(out / "wu2022_table_s1_rows.csv", rows, fields)
        source_taxa.update(taxa); source_parts.append("publisher_Table_S1")
        table_sheet = sheet; table_rows_n = len(rows)
    sra_rows_n = 0; sra_taxa_n = 0
    if args.sra_runinfo and args.sra_runinfo.exists():
        rows, taxa = extract_sra_taxa(args.sra_runinfo)
        write_csv(out / "prjna665925_run_taxa.csv", rows)
        source_taxa.update(taxa); source_parts.append("PRJNA665925_RunInfo")
        sra_rows_n = len(rows); sra_taxa_n = len(taxa)
    source_taxa = sorted(source_taxa)

    sess = requests.Session(); sess.headers.update({"User-Agent": "Mozilla/5.0 chun-public-backbone-audit/0.3"})
    catalog = fetch_json(sess, args.tpia_base.rstrip("/") + "/selectAllassemblies")
    if isinstance(catalog, dict):
        for key in ("data", "rows", "result"):
            if isinstance(catalog.get(key), list): catalog = catalog[key]; break
    if not isinstance(catalog, list): raise SystemExit(f"Unexpected selectAllassemblies payload type: {type(catalog)}")

    catrows = []; bytax = defaultdict(list)
    for r in catalog:
        name = clean(r.get("name")); tax = display_taxon(name); zip_url = ""
        if r.get("hasZipFile"):
            zip_url = args.tpia_base.rstrip("/") + "/web/All_assemblies/Fasta/" + quote(f"{r.get('ID')}_{name}.zip", safe="_")
        row = {k: clean(v) for k, v in r.items()}
        row.update({"normalized_taxon": tax, "assembly_zip_url": zip_url})
        catrows.append(row)
        if tax: bytax[tax].append(row)
    write_csv(out / "tpia_allassemblies_catalog.csv", catrows)

    bulk = fetch_json(sess, args.tpia_base.rstrip("/") + "/selectdownload_data", params={"Type": "Transcriptome data"})
    if isinstance(bulk, dict):
        for key in ("data", "rows", "result"):
            if isinstance(bulk.get(key), list): bulk = bulk[key]; break
    bulk_rows = []; bulk_by_tax = defaultdict(list)
    if isinstance(bulk, list):
        for r in bulk:
            row = {k: clean(v) for k, v in r.items()}
            tax = bulk_assembly_taxon(row.get("description", ""), row.get("fileName", ""))
            is_assembly = row.get("description", "").lower().startswith("transcriptome assembly of camellia")
            url = ""
            if is_assembly and row.get("fileName"):
                url = args.tpia_base.rstrip("/") + "/web/Download/Transcriptome_data/" + quote(row["fileName"])
            row.update({"normalized_taxon": tax, "is_transcriptome_assembly": str(bool(is_assembly)), "bulk_download_url": url})
            bulk_rows.append(row)
            if tax and is_assembly: bulk_by_tax[tax].append(row)
        write_csv(out / "tpia_transcriptome_bulk_catalog.csv", bulk_rows)
    else:
        (out / "tpia_transcriptome_bulk_raw.json").write_text(json.dumps(bulk, ensure_ascii=False, indent=2) + "\n")

    alias_rows = []
    cross = []; unresolved = []
    preferred = []
    for source_tax in source_taxa:
        query_tax = source_tax
        match_basis = "exact_archive_tpia_species_name"
        if source_tax in ARCHIVE_TPIA_CROSSWALK:
            query_tax, match_basis = ARCHIVE_TPIA_CROSSWALK[source_tax]
            alias_rows.append({"source_archive_taxon": source_tax, "tpia_resource_taxon": query_tax, "match_basis": match_basis, "scope": "Wu2022_PRJNA665925_to_TPIA_only"})
        cat_hits = bytax.get(query_tax, [])
        cat_zip_hits = [h for h in cat_hits if h.get("assembly_zip_url")]
        bulk_hits = bulk_by_tax.get(query_tax, [])
        # Prefer bulk catalog because it exposes explicit file name + advertised size;
        # otherwise use the live all-assemblies ZIP URL.
        if bulk_hits:
            choice = sorted(bulk_hits, key=lambda x: (0 if x.get("sourceData") == "PRJNA665925" else 1, x.get("no", "")))[0]
            preferred_url = choice.get("bulk_download_url", "")
            preferred_name = choice.get("fileName", "")
            preferred_size = choice.get("size", "")
            preferred_source = "tpia_bulk_transcriptome_catalog"
        elif cat_zip_hits:
            choice = sorted(cat_zip_hits, key=lambda x: (0 if x.get("sourceData") == "PRJNA665925" else 1, x.get("ID", "")))[0]
            preferred_url = choice.get("assembly_zip_url", "")
            preferred_name = f"{choice.get('ID')}_{choice.get('name')}.zip"
            preferred_size = ""
            preferred_source = "tpia_selectAllassemblies"
        else:
            preferred_url = preferred_name = preferred_size = preferred_source = ""
            unresolved.append(source_tax)

        cross.append({
            "source_taxon": source_tax,
            "tpia_resource_taxon": query_tax,
            "match_basis": match_basis,
            "tpia_catalog_match_count": len(cat_hits),
            "tpia_catalog_zip_count": len(cat_zip_hits),
            "tpia_bulk_assembly_count": len(bulk_hits),
            "preferred_assembly_source": preferred_source,
            "preferred_assembly_file": preferred_name,
            "preferred_assembly_size_advertised": preferred_size,
            "preferred_assembly_url": preferred_url,
            "status": "downloadable_assembly" if preferred_url else "unresolved_no_public_assembly",
        })
        if preferred_url:
            preferred.append({
                "source_taxon": source_tax,
                "resource_taxon": query_tax,
                "match_basis": match_basis,
                "assembly_source": preferred_source,
                "assembly_file": preferred_name,
                "advertised_size": preferred_size,
                "assembly_url": preferred_url,
                "analysis_role": "species_level_nuclear_backbone_input",
                "claim_ceiling": "Wu2022 resource crosswalk; one assembly per species-level SRA taxon; not exact reconstruction of all 116 paper tips",
            })
    write_csv(out / "wu2022_archive_tpia_alias_crosswalk.csv", alias_rows, ["source_archive_taxon","tpia_resource_taxon","match_basis","scope"])
    write_csv(out / "wu2022_tpia_species_crosswalk.csv", cross)
    write_csv(out / "wu2022_preferred_assembly_manifest.csv", preferred)

    summary = {
        "source": "+".join(source_parts),
        "doi": "10.1111/tpj.15799",
        "publisher_table_s1_recovered": bool(args.supp_xlsx and args.supp_xlsx.exists()),
        "table_s1_sheet": table_sheet,
        "table_s1_rows": table_rows_n,
        "sra_runinfo_rows": sra_rows_n,
        "sra_unique_camellia_species_level_taxa": sra_taxa_n,
        "source_unique_camellia_species_level_taxa": len(source_taxa),
        "tpia_catalog_rows": len(catrows),
        "tpia_unique_camellia_species_level_taxa": len(bytax),
        "tpia_bulk_rows": len(bulk_rows),
        "tpia_bulk_transcriptome_assemblies": sum(1 for r in bulk_rows if r.get("is_transcriptome_assembly") == "True"),
        "archive_tpia_explicit_crosswalk_rows": len(alias_rows),
        "source_taxa_with_downloadable_preferred_assembly": len(preferred),
        "unresolved_source_taxa": unresolved,
        "coverage_fraction": len(preferred) / len(source_taxa) if source_taxa else None,
        "claim_ceiling": "asset/crosswalk gate only; one assembly per PRJNA665925 species-level taxon; duplicated cultivated-tea paper tips omitted unless separately crosswalked; no topology or transition inference",
    }
    (out / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))

if __name__ == "__main__": main()
