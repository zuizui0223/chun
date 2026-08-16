#!/usr/bin/env python3
"""Audit Fan et al. 2026 Data S1 without assuming its schema.

Writes sheet inventory, exact column names, row counts, and CSV exports.  If a
sheet contains taxon/flower-colour-like columns they are flagged for downstream
manual/scientific admission; this script does not invent mappings.
"""
from __future__ import annotations
import argparse, csv, json, pathlib, re
from openpyxl import load_workbook

KEY_RE = re.compile(r"(species|taxon|name|flower|petal|colo[u]?r|section|country|region|location|accession)", re.I)

def norm(v):
    if v is None: return ""
    return str(v).strip()

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("xlsx", type=pathlib.Path)
    ap.add_argument("--out-dir", type=pathlib.Path, required=True)
    a=ap.parse_args(); a.out_dir.mkdir(parents=True, exist_ok=True)
    wb=load_workbook(a.xlsx, read_only=True, data_only=True)
    inventory=[]
    for ws in wb.worksheets:
        rows=list(ws.iter_rows(values_only=True))
        nonempty=[r for r in rows if any(norm(x) for x in r)]
        if not nonempty:
            inventory.append({"sheet":ws.title,"n_nonempty_rows":0,"header_row":"","columns":"","candidate_key_columns":""})
            continue
        # Find first row with at least two non-empty cells; do not assume row 1.
        hi=next(i for i,r in enumerate(nonempty) if sum(bool(norm(x)) for x in r)>=2)
        header=[norm(x) or f"unnamed_{j+1}" for j,x in enumerate(nonempty[hi])]
        candidate=[h for h in header if KEY_RE.search(h)]
        data=[]
        for r in nonempty[hi+1:]:
            rr=list(r)+[None]*(len(header)-len(r))
            data.append({header[j]: norm(rr[j]) for j in range(len(header))})
        safe=re.sub(r"[^A-Za-z0-9_.-]+","_",ws.title).strip("_") or "sheet"
        out=a.out_dir/f"{safe}.csv"
        with out.open("w",newline="",encoding="utf-8") as f:
            w=csv.DictWriter(f,fieldnames=header); w.writeheader(); w.writerows(data)
        inventory.append({
            "sheet":ws.title,"n_nonempty_rows":len(nonempty),"header_row":hi+1,
            "n_data_rows":len(data),"columns":" | ".join(header),
            "candidate_key_columns":" | ".join(candidate),"csv":str(out)
        })
    with (a.out_dir/"sheet_inventory.csv").open("w",newline="",encoding="utf-8") as f:
        fields=list(inventory[0]) if inventory else ["sheet"]
        w=csv.DictWriter(f,fieldnames=fields); w.writeheader(); w.writerows(inventory)
    (a.out_dir/"audit_summary.json").write_text(json.dumps({"sheets":inventory},ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    print(json.dumps({"sheets":inventory},ensure_ascii=False,indent=2))
if __name__=="__main__": main()
